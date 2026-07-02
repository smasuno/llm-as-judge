"""
perform_writeup.py — output stage for the CSRP pipeline (build-order step 9).

Three deliverables from one scoring `result`:
  (a) Markdown report (EN + JA) — LLM-authored section-by-section, mirroring
      ../AI-Scientist/ai_scientist/perform_writeup.py: a `per_section_tips` dict,
      a system prompt, a per-section prompt, an optional refinement pass, and a
      shared `msg_history` so each section sees the prior ones. Our analogue of
      the original's aider `coder` is `get_response_from_llm`.
  (b) JSON (EN + JA) — deterministic, the machine-readable record.
  (c) Excel workbook — deterministic, formatted, DCF-ready, with an "English"
      and a "日本語" tab: recommended CSRP value + component scores at the top,
      detailed rationale + citations at the bottom.

The Excel/JSON are deterministic (free to test); only the markdown costs tokens.
"""

import json
import os
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from .llm import get_response_from_llm

DEFAULT_MODEL: str = "claude-sonnet-4-6"

# --- Prompts (markdown authoring; max 5) -------------------------------------

WRITEUP_SYSTEM_PROMPT: str = (
    "You are a valuation analyst writing a Company-Specific Risk Premium (CSRP) "
    "report for an M&A engagement. Use ONLY the data provided — scores, values, "
    "rationales, and citations. Quote the Japanese citation passages where they "
    "support a point. Never invent numbers or evidence. Write in clean Markdown."
)

per_section_tips: dict[str, str] = {
    "Executive Summary": (
        "- State the final recommended CSRP premium (bps range + midpoint) and the composite score.\n"
        "- One paragraph: the overall risk picture and what drives it."
    ),
    "Per-Category Scorecard": (
        "- A table of the 8 weighted categories with their scores and weights.\n"
        "- Under each category, bullet the notable signals with score, value, and a cited reason."
    ),
    "Japan-Specific Supplemental Flags": (
        "- Summarize the JPN supplemental signals (政策保有株式, メインバンク, etc.).\n"
        "- These are reported for context and are NOT in the weighted composite."
    ),
    "Top 3 Risk Drivers": (
        "- The three highest-scored signals. For each: what it is, the value, and the cited evidence."
    ),
    "Methodology Note": (
        "- Comparables mode (off), the model used, how many signals were unscored and why,\n"
        "- and that scores map to bps via the taxonomy's score→premium table."
    ),
}

SECTION_PROMPT: str = (
    "Write the **{section}** section of the CSRP report for {company}, in {language}.\n"
    "Tips:\n{tips}\n\nUse the report data already provided. Cite Japanese passages where relevant."
)

REFINEMENT_PROMPT: str = (
    "Refine only the {section} section you just wrote: tighten it, fix any errors, "
    "remove redundancy, and ensure every figure is grounded in the provided data. "
    "Keep it in the same language."
)


# --- shared helpers ----------------------------------------------------------


def _slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_") or "company"


def _t(lang: str, en: str, ja: str) -> str:
    return en if lang == "en" else ja


def _signal_index(taxonomy: dict[str, Any]) -> dict[str, dict[str, Any]]:
    """sid -> {"def": signal_dict, "cat": category_dict}."""
    idx: dict[str, dict[str, Any]] = {}
    for cat in taxonomy["categories"]:
        for sig in cat["signals"]:
            idx[sig["id"]] = {"def": sig, "cat": cat}
    return idx


def _citation_text(citation: dict[str, Any] | None) -> str:
    """Flatten a citation (NLP or structured) into one readable string."""
    if not citation:
        return ""
    if citation.get("citation_type") == "nlp":
        return f"【{citation.get('section')}】{citation.get('passage', '')}"
    parts = [
        f"{it.get('line_item')}={it.get('value')}（{it.get('section')}）"
        for it in citation.get("inputs", [])
    ]
    return " / ".join(parts)


def recommended_bps(result: dict[str, Any]) -> int | None:
    """Midpoint of the bps range — the single number an analyst adds to WACC."""
    lo, hi = result.get("csrp_bps_low"), result.get("csrp_bps_high")
    if lo is None or hi is None:
        return None
    return round((lo + hi) / 2)


# =============================================================================
# (a) Markdown — LLM-authored, section-by-section (mirrors AI-Scientist)
# =============================================================================


def _format_context(result: dict[str, Any], taxonomy: dict[str, Any],
                    metadata: dict[str, Any]) -> str:
    """Compact grounding block fed once to the writer (analogue of notes.txt)."""
    idx = _signal_index(taxonomy)
    lines = [
        f"Company: {metadata.get('company')}  |  Filing: {metadata.get('filing_date')}  "
        f"|  Run: {metadata.get('run_date')}  |  Model: {metadata.get('model')}",
        f"Composite score: {result['composite_score']}  ->  "
        f"{result['csrp_bps_low']}-{result['csrp_bps_high']} bps "
        f"(recommended midpoint {recommended_bps(result)} bps)",
        f"comparables_mode: {result.get('comparables_mode')}",
        "",
        "Category scores: " + ", ".join(
            f"{cid}={sc}" for cid, sc in result["category_scores"].items()),
        "",
        "Signals (id | name | score | scored_by | value | rationale | citation):",
    ]
    for sid, r in result["signals"].items():
        if r["status"] != "scored":
            continue
        name = idx.get(sid, {}).get("def", {}).get("name_en", sid)
        lines.append(
            f"- {sid} | {name} | {r['score']} | {r['scored_by']} | {r['value']} "
            f"| {r.get('rationale_en')} | {_citation_text(r['citation'])[:160]}"
        )
    unscored = result.get("unscored_signals", [])
    lines.append(f"\nUnscored signals ({len(unscored)}): "
                 + ", ".join(f"{u['id']} ({u['reason']})" for u in unscored[:12]))
    return "\n".join(lines)


def write_markdown(result: dict[str, Any], taxonomy: dict[str, Any],
                   metadata: dict[str, Any], client: Any, model: str = DEFAULT_MODEL,
                   language: str = "English", num_refinements: int = 0) -> str:
    """Author the report section-by-section in `language`, sharing msg_history."""
    context = _format_context(result, taxonomy, metadata)
    msg_history: list[Any] = []
    parts: list[str] = []
    for i, (section, tips) in enumerate(per_section_tips.items()):
        prompt = SECTION_PROMPT.format(section=section, company=metadata.get("company"),
                                       language=language, tips=tips)
        if i == 0:  # provide the grounding data once, up front
            prompt = f"Here is the CSRP report data:\n\n{context}\n\n" + prompt
        text, msg_history = get_response_from_llm(
            prompt, client=client, model=model,
            system_message=WRITEUP_SYSTEM_PROMPT, msg_history=msg_history, temperature=0.3,
        )
        for _ in range(num_refinements):
            text, msg_history = get_response_from_llm(
                REFINEMENT_PROMPT.format(section=section), client=client, model=model,
                system_message=WRITEUP_SYSTEM_PROMPT, msg_history=msg_history, temperature=0.3,
            )
        parts.append(text.strip())
    return "\n\n".join(parts)


# =============================================================================
# (b) JSON — deterministic
# =============================================================================


def build_json(result: dict[str, Any], taxonomy: dict[str, Any],
               metadata: dict[str, Any], lang: str) -> dict[str, Any]:
    idx = _signal_index(taxonomy)
    nk = "name_en" if lang == "en" else "name_ja"
    rk = "rationale_en" if lang == "en" else "rationale_ja"
    signals: dict[str, Any] = {}
    for sid, r in result["signals"].items():
        signals[sid] = {
            "name": idx.get(sid, {}).get("def", {}).get(nk, sid),
            "score": r["score"], "value": r["value"], "flag": r["flag"],
            "scored_by": r["scored_by"], "rationale": r.get(rk),
            "citation": r["citation"], "status": r["status"],
        }
    return {
        "metadata": metadata,
        "signals": signals,
        "category_scores": result["category_scores"],
        "composite_score": result["composite_score"],
        "csrp_bps_low": result["csrp_bps_low"],
        "csrp_bps_high": result["csrp_bps_high"],
        "recommended_bps": recommended_bps(result),
        "japan_supplemental": result["japan_supplemental"],
        "unscored_signals": result["unscored_signals"],
        "search_calls": [],
    }


# =============================================================================
# (c) Excel — deterministic, formatted, DCF-ready
# =============================================================================

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")      # dark navy
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14, color="1F3864")
_DCF_FILL = PatternFill("solid", fgColor="FFF2CC")          # highlight
_DCF_FONT = Font(bold=True, size=12, color="9C5700")
_SECTION_FONT = Font(bold=True, size=11, color="1F3864")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _hrow(ws, row: int, headers: list[str]) -> None:
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.fill, cell.font, cell.border = _HEADER_FILL, _HEADER_FONT, _BORDER


def _fill_sheet(ws, lang: str, result: dict[str, Any], taxonomy: dict[str, Any],
                metadata: dict[str, Any]) -> None:
    idx = _signal_index(taxonomy)
    nk = "name_en" if lang == "en" else "name_ja"
    rk = "rationale_en" if lang == "en" else "rationale_ja"
    widths = [22, 30, 10, 12, 22, 46, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title + metadata
    ws.merge_cells("A1:G1")
    ws["A1"] = _t(lang, "Company-Specific Risk Premium (CSRP)",
                  "会社固有リスクプレミアム（CSRP）") + f" — {metadata.get('company')}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = _t(lang, f"Filing: {metadata.get('filing_date')}   Run: {metadata.get('run_date')}"
                       f"   Model: {metadata.get('model')}   Comparables: {metadata.get('comparables_mode')}",
                  f"提出: {metadata.get('filing_date')}   実行: {metadata.get('run_date')}"
                  f"   モデル: {metadata.get('model')}   同業比較: {metadata.get('comparables_mode')}")

    # DCF INPUT headline
    mid = recommended_bps(result)
    lo, hi = result.get("csrp_bps_low"), result.get("csrp_bps_high")
    r = 4
    ws.cell(row=r, column=1, value=_t(lang, "DCF INPUT — Recommended CSRP premium (add to WACC)",
                                      "DCF入力 — 推奨CSRPプレミアム（WACCに加算）")).font = _DCF_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    bps_cell = ws.cell(row=r, column=5, value=(mid if mid is not None else "N/A"))
    bps_cell.fill, bps_cell.font = _DCF_FILL, _DCF_FONT
    if mid is not None:
        bps_cell.number_format = '0 "bps"'
        pct = ws.cell(row=r, column=6, value=mid / 10000)
        pct.fill, pct.font, pct.number_format = _DCF_FILL, _DCF_FONT, "0.00%"
        ws.cell(row=r + 1, column=5,
                value=_t(lang, f"range {lo}–{hi} bps", f"レンジ {lo}–{hi} bps"))
    ws.cell(row=r + 1, column=1,
            value=_t(lang, f"Composite score: {result['composite_score']}",
                     f"総合スコア: {result['composite_score']}"))

    # Component values table
    r += 3
    ws.cell(row=r, column=1, value=_t(lang, "Component values", "構成要素スコア")).font = _SECTION_FONT
    r += 1
    _hrow(ws, r, [_t(lang, "Category", "カテゴリー"), _t(lang, "Name", "名称"),
                  _t(lang, "Weight", "ウェイト"), _t(lang, "Score (1–5)", "スコア(1–5)")])
    score_start = r + 1
    for cat in taxonomy["categories"]:
        if cat.get("supplemental"):
            continue
        r += 1
        ws.cell(row=r, column=1, value=cat["id"]).border = _BORDER
        ws.cell(row=r, column=2, value=cat[nk]).border = _BORDER
        wcell = ws.cell(row=r, column=3, value=cat["default_weight"]); wcell.border = _BORDER; wcell.number_format = "0.00"
        scell = ws.cell(row=r, column=4, value=result["category_scores"].get(cat["id"]))
        scell.border = _BORDER; scell.number_format = "0.00"
    score_end = r
    # composite row
    r += 1
    ws.cell(row=r, column=2, value=_t(lang, "COMPOSITE", "総合")).font = Font(bold=True)
    cc = ws.cell(row=r, column=4, value=result["composite_score"]); cc.font = Font(bold=True); cc.number_format = "0.00"
    if score_end >= score_start:
        ws.conditional_formatting.add(
            f"D{score_start}:D{score_end}",
            ColorScaleRule(start_type="num", start_value=1, start_color="63BE7B",
                           mid_type="num", mid_value=3, mid_color="FFEB84",
                           end_type="num", end_value=5, end_color="F8696B"))

    # Japan supplemental (excluded from composite)
    r += 2
    ws.cell(row=r, column=1, value=_t(lang, "Japan supplemental (excluded from composite)",
                                      "日本固有の補足（総合スコア対象外）")).font = _SECTION_FONT
    for cat in taxonomy["categories"]:
        if not cat.get("supplemental"):
            continue
        sup = result.get("japan_supplemental", {}).get(cat["id"], {})
        sup_sigs = sup.get("signals", {})
        for sig in cat["signals"]:
            r += 1
            ws.cell(row=r, column=1, value=sig["id"]).border = _BORDER
            ws.cell(row=r, column=2, value=sig[nk]).border = _BORDER
            sc = sup_sigs.get(sig["id"], {}).get("score")
            ws.cell(row=r, column=4, value=sc).border = _BORDER

    # Rationale & citations (bottom, detailed)
    r += 2
    ws.cell(row=r, column=1, value=_t(lang, "Rationale & citations", "根拠と出典")).font = _SECTION_FONT
    r += 1
    _hrow(ws, r, [_t(lang, "Category", "カテゴリー"), _t(lang, "Signal", "シグナル"),
                  _t(lang, "Score", "スコア"), _t(lang, "Scored by", "採点方法"),
                  _t(lang, "Value", "値"), _t(lang, "Rationale", "根拠"),
                  _t(lang, "Citation", "出典")])
    for cat in taxonomy["categories"]:
        for sig in cat["signals"]:
            sid = sig["id"]
            rr = result["signals"].get(sid, {})
            r += 1
            ws.cell(row=r, column=1, value=cat["id"]).border = _BORDER
            ws.cell(row=r, column=2, value=sig[nk]).border = _BORDER
            ws.cell(row=r, column=3, value=rr.get("score")).border = _BORDER
            ws.cell(row=r, column=4, value=rr.get("scored_by")).border = _BORDER
            ws.cell(row=r, column=5, value=str(rr.get("value")) if rr.get("value") is not None else "").border = _BORDER
            rat = ws.cell(row=r, column=6, value=rr.get(rk)); rat.border = _BORDER; rat.alignment = _WRAP
            cit = ws.cell(row=r, column=7, value=_citation_text(rr.get("citation"))); cit.border = _BORDER; cit.alignment = _WRAP

    ws.freeze_panes = "A4"


def write_excel(result: dict[str, Any], taxonomy: dict[str, Any],
                metadata: dict[str, Any], out_path: str) -> str:
    """Write the DCF-ready workbook with English + 日本語 tabs."""
    wb = Workbook()
    ws_en = wb.active
    ws_en.title = "English"
    _fill_sheet(ws_en, "en", result, taxonomy, metadata)
    ws_ja = wb.create_sheet("日本語")
    _fill_sheet(ws_ja, "ja", result, taxonomy, metadata)
    wb.save(out_path)
    return out_path


# =============================================================================
# Orchestration
# =============================================================================


def perform_writeup(result: dict[str, Any], taxonomy: dict[str, Any],
                    metadata: dict[str, Any], client: Any = None,
                    model: str = DEFAULT_MODEL, out_dir: str = "outputs",
                    num_refinements: int = 0) -> dict[str, str]:
    """Write JSON + Excel (always) and Markdown (when a client is given)."""
    slug = _slug(metadata.get("company", "company"))
    folder = os.path.join(out_dir, slug)
    os.makedirs(folder, exist_ok=True)
    paths: dict[str, str] = {}

    for lang in ("en", "ja"):
        p = os.path.join(folder, f"csrp_{slug}_{lang}.json")
        with open(p, "w", encoding="utf-8") as f:
            json.dump(build_json(result, taxonomy, metadata, lang), f, ensure_ascii=False, indent=2)
        paths[f"json_{lang}"] = p

    paths["xlsx"] = write_excel(result, taxonomy, metadata,
                                os.path.join(folder, f"csrp_{slug}.xlsx"))

    if client is not None:
        for lang, name in (("en", "English"), ("ja", "Japanese")):
            md = write_markdown(result, taxonomy, metadata, client, model, name, num_refinements)
            p = os.path.join(folder, f"csrp_{slug}_{lang}.md")
            with open(p, "w", encoding="utf-8") as f:
                f.write(md)
            paths[f"md_{lang}"] = p
    return paths


# --- Smoke test (no API by default) ------------------------------------------

if __name__ == "__main__":
    import sys
    from .perform_review import score_signals, aggregate

    here = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
    taxonomy = json.load(open(os.path.join(here, "taxonomy.json")))["csrp_framework"]

    # Synthetic records (structured threshold + structured informational) — no API.
    def struct(flag: bool | None, value: Any, en: str, ja: str) -> dict[str, Any]:
        return {"value": value, "flag": flag, "evidence_en": en, "evidence_ja": ja,
                "citation": {"citation_type": "structured", "doc": "SAAF.pdf",
                             "inputs": [{"line_item": "short_term_debt", "value": 4530000,
                                         "section": "借入金等明細表", "passage": "短期借入金 4,530,000"}]},
                "layer": "structured", "status": "scored"}

    records = {
        "FIN_01": struct(True, 6.9, "Net Debt/EBITDA ≈ 6.9x (>4.0x).", "ネット有利子負債/EBITDA ≈ 6.9倍（>4.0倍）。"),
        "FIN_02": struct(False, 6.73, "EBITDA/Interest ≈ 6.73x.", "EBITDA/支払利息 ≈ 6.73倍。"),
        "FIN_03": struct(True, "57% ≤24m", "57% of debt due ≤24m.", "有利子負債の57%が24か月以内。"),
    }
    scored = score_signals(taxonomy, records, client=None)
    result = aggregate(taxonomy, scored)
    result["signals"] = scored
    result["comparables_mode"] = False

    metadata = {"company": "SAAF Co., Ltd.", "filing_date": "2025-03-31",
                "run_date": "2026-06-06", "model": DEFAULT_MODEL, "comparables_mode": False}

    out_dir = os.path.join(here, "outputs", "_smoke")
    paths = perform_writeup(result, taxonomy, metadata, client=None, out_dir=out_dir)
    print("wrote:", {k: os.path.relpath(v, here) for k, v in paths.items()})

    # Reopen the workbook and assert structure (no API).
    from openpyxl import load_workbook
    wb = load_workbook(paths["xlsx"])
    print("sheets:", wb.sheetnames)
    en = wb["English"]
    print("recommended bps cell (E4):", en["E4"].value, "| % cell (F4):", en["F4"].value)
    print("composite from json_en:", json.load(open(paths["json_en"]))["recommended_bps"], "bps midpoint")

    if "--llm" in sys.argv:
        from .llm import create_client
        client, model = create_client(DEFAULT_MODEL)
        paths = perform_writeup(result, taxonomy, metadata, client=client, model=model, out_dir=out_dir)
        print("with markdown:", {k: os.path.relpath(v, here) for k, v in paths.items() if k.startswith("md")})
